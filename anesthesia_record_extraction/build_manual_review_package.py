from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
MONITOR_DIR = ROOT / "批量监测CSV"
AXIS_DIR = ROOT / "批量坐标对齐图"
REVIEW_DIR = ROOT / "人工核对包"
REVIEW_CSV = REVIEW_DIR / "人工核对清单.csv"
REVIEW_XLSX = REVIEW_DIR / "人工核对清单.xlsx"
README_TXT = REVIEW_DIR / "说明.txt"


def open_with_locked_copy(path: Path, mode: str, **kwargs):
    try:
        handle = path.open(mode, **kwargs)
        return handle, path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}__locked_copy{path.suffix}")
        handle = fallback.open(mode, **kwargs)
        return handle, fallback


def read_summary(path: Path, key_name: str, value_name: str) -> dict[str, str]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return {row[key_name]: row[value_name] for row in reader if row.get(key_name)}


def gather_cases() -> list[dict[str, object]]:
    monitor_status = read_summary(MONITOR_DIR / "处理汇总.csv", "case_name", "status")
    axis_status = read_summary(AXIS_DIR / "对齐图汇总.csv", "case_name", "status")

    case_names = set(monitor_status) | set(axis_status)

    for path in MONITOR_DIR.glob("*.csv"):
        if path.name in {"处理汇总.csv"} or path.stem.endswith("__locked_copy"):
            continue
        case_names.add(path.stem)

    rows: list[dict[str, object]] = []
    for case_name in sorted(case_names):
        csv_path = MONITOR_DIR / f"{case_name}.csv"
        overlay_images = sorted(MONITOR_DIR.glob(f"{case_name}_overlay_page*.png"))
        axis_images = sorted(AXIS_DIR.glob(f"{case_name}_axis_overlay_p*.png"))
        row = {
            "case_name": case_name,
            "monitor_status": monitor_status.get(case_name, ""),
            "axis_status": axis_status.get(case_name, ""),
            "pdf_path": "",
            "csv_path": str(csv_path) if csv_path.exists() else "",
            "overlay_images": [str(path) for path in overlay_images],
            "axis_images": [str(path) for path in axis_images],
            "review_csv": "",
            "review_axis": "",
            "review_time": "",
            "issue_notes": "",
            "final_decision": "",
        }
        rows.append(row)

    for row in rows:
        case_name = str(row["case_name"])
        pdf_candidates = list((ROOT / "阳性").glob(f"{case_name}/*.pdf")) + list((ROOT / "正常").glob(f"{case_name}/*.pdf"))
        if pdf_candidates:
            row["pdf_path"] = str(pdf_candidates[0])
    return rows


def write_review_csv(rows: list[dict[str, object]]) -> None:
    fieldnames = [
        "case_name",
        "monitor_status",
        "axis_status",
        "pdf_path",
        "csv_path",
        "overlay_page1",
        "overlay_page2",
        "overlay_page3",
        "axis_page1",
        "axis_page2",
        "axis_page3",
        "review_csv",
        "review_axis",
        "review_time",
        "issue_notes",
        "final_decision",
    ]
    f, _used_path = open_with_locked_copy(REVIEW_CSV, "w", encoding="utf-8-sig", newline="")
    with f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            overlay_images = row["overlay_images"]
            axis_images = row["axis_images"]
            writer.writerow(
                {
                    "case_name": row["case_name"],
                    "monitor_status": row["monitor_status"],
                    "axis_status": row["axis_status"],
                    "pdf_path": row["pdf_path"],
                    "csv_path": row["csv_path"],
                    "overlay_page1": overlay_images[0] if len(overlay_images) >= 1 else "",
                    "overlay_page2": overlay_images[1] if len(overlay_images) >= 2 else "",
                    "overlay_page3": overlay_images[2] if len(overlay_images) >= 3 else "",
                    "axis_page1": axis_images[0] if len(axis_images) >= 1 else "",
                    "axis_page2": axis_images[1] if len(axis_images) >= 2 else "",
                    "axis_page3": axis_images[2] if len(axis_images) >= 3 else "",
                    "review_csv": "",
                    "review_axis": "",
                    "review_time": "",
                    "issue_notes": "",
                    "final_decision": "",
                }
            )


def set_link(cell, target: str, label: str) -> None:
    if not target:
        cell.value = ""
        return
    cell.value = label
    cell.hyperlink = target
    cell.style = "Hyperlink"


def write_review_workbook(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "人工核对清单"

    headers = [
        "病例",
        "监测CSV状态",
        "坐标对齐图状态",
        "原始PDF",
        "CSV",
        "监测叠图1",
        "监测叠图2",
        "监测叠图3",
        "对齐图1",
        "对齐图2",
        "对齐图3",
        "CSV核对",
        "图像核对",
        "时间核对",
        "问题记录",
        "最终结论",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row_idx, row in enumerate(rows, start=2):
        overlay_images = row["overlay_images"]
        axis_images = row["axis_images"]
        ws.cell(row=row_idx, column=1, value=row["case_name"])
        ws.cell(row=row_idx, column=2, value=row["monitor_status"])
        ws.cell(row=row_idx, column=3, value=row["axis_status"])
        set_link(ws.cell(row=row_idx, column=4), row["pdf_path"], "打开PDF")
        set_link(ws.cell(row=row_idx, column=5), row["csv_path"], "打开CSV")
        set_link(ws.cell(row=row_idx, column=6), overlay_images[0] if len(overlay_images) >= 1 else "", "打开叠图1")
        set_link(ws.cell(row=row_idx, column=7), overlay_images[1] if len(overlay_images) >= 2 else "", "打开叠图2")
        set_link(ws.cell(row=row_idx, column=8), overlay_images[2] if len(overlay_images) >= 3 else "", "打开叠图3")
        set_link(ws.cell(row=row_idx, column=9), axis_images[0] if len(axis_images) >= 1 else "", "打开对齐图1")
        set_link(ws.cell(row=row_idx, column=10), axis_images[1] if len(axis_images) >= 2 else "", "打开对齐图2")
        set_link(ws.cell(row=row_idx, column=11), axis_images[2] if len(axis_images) >= 3 else "", "打开对齐图3")
        for col in range(12, 17):
            ws.cell(row=row_idx, column=col, value="")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        1: 20,
        2: 16,
        3: 16,
        4: 12,
        5: 12,
        6: 12,
        7: 12,
        8: 12,
        9: 12,
        10: 12,
        11: 12,
        12: 12,
        13: 12,
        14: 12,
        15: 36,
        16: 16,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    try:
        wb.save(REVIEW_XLSX)
    except PermissionError:
        wb.save(REVIEW_XLSX.with_name(f"{REVIEW_XLSX.stem}__locked_copy{REVIEW_XLSX.suffix}"))


def write_readme() -> None:
    lines = [
        "人工核对包说明",
        "",
        "1. 先打开“人工核对清单.xlsx”。",
        "2. 每行对应 1 个病例，点击“打开CSV / 打开叠图 / 打开对齐图”进入对应文件。",
        "3. 建议核对顺序：CSV -> 监测叠图 -> 坐标对齐图 -> 时间轴。",
        "4. 可在 Excel 里直接填写“CSV核对、图像核对、时间核对、问题记录、最终结论”。",
        "5. 如果某例状态不是 ok，优先看状态列和问题记录。",
    ]
    README_TXT.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build CSV/XLSX manual review files from extraction outputs.")
    parser.add_argument(
        "--workspace",
        type=Path,
        default=ROOT,
        help="Directory containing extraction outputs and case groups. Defaults to this script directory.",
    )
    return parser.parse_args()


def configure_paths(workspace: Path) -> None:
    global ROOT, MONITOR_DIR, AXIS_DIR, REVIEW_DIR, REVIEW_CSV, REVIEW_XLSX, README_TXT
    ROOT = workspace
    MONITOR_DIR = ROOT / "批量监测CSV"
    AXIS_DIR = ROOT / "批量坐标对齐图"
    REVIEW_DIR = ROOT / "人工核对包"
    REVIEW_CSV = REVIEW_DIR / "人工核对清单.csv"
    REVIEW_XLSX = REVIEW_DIR / "人工核对清单.xlsx"
    README_TXT = REVIEW_DIR / "说明.txt"


def main() -> None:
    args = parse_args()
    configure_paths(args.workspace.resolve())
    REVIEW_DIR.mkdir(exist_ok=True)
    rows = gather_cases()
    write_review_csv(rows)
    write_review_workbook(rows)
    write_readme()
    print(f"已生成: {REVIEW_XLSX}")


if __name__ == "__main__":
    main()
