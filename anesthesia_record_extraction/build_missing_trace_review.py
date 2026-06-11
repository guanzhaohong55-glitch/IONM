from __future__ import annotations

import argparse
import csv
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
MONITOR_DIR = ROOT / "批量监测CSV"
AXIS_DIR = ROOT / "批量坐标对齐图"
OUTPUT_DIR = ROOT / "人工核对包"
OUTPUT_XLSX = OUTPUT_DIR / "优先补录清单_missing_trace_cluster.xlsx"


def find_pdf(case_name: str) -> str:
    for group_name in ("阳性", "正常"):
        group_dir = ROOT / group_name / case_name
        if not group_dir.exists():
            continue
        pdfs = sorted(group_dir.glob("*.pdf"))
        if pdfs:
            return str(pdfs[0])
    return ""


def read_csv_text_with_fallback(path: Path) -> str:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    return path.read_text(encoding="utf-8-sig")


def collect_missing_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for csv_path in sorted(MONITOR_DIR.glob("*.csv")):
        if csv_path.name == "处理汇总.csv" or csv_path.stem.endswith("__locked_copy"):
            continue
        case_name = csv_path.stem
        pdf_path = find_pdf(case_name)
        text = read_csv_text_with_fallback(csv_path)
        with io.StringIO(text) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if (row.get("note") or "").strip() != "missing trace cluster":
                    continue
                page = (row.get("page") or "").strip()
                overlay_path = MONITOR_DIR / f"{case_name}_overlay_page{page}.png" if page else None
                axis_path = AXIS_DIR / f"{case_name}_axis_overlay_p{page}.png" if page else None
                rows.append(
                    {
                        "case_name": case_name,
                        "time": (row.get("time") or "").strip(),
                        "page": page,
                        "bp_source": (row.get("bp_source") or "").strip(),
                        "confidence": (row.get("confidence") or "").strip(),
                        "note": (row.get("note") or "").strip(),
                        "csv_path": str(csv_path),
                        "overlay_path": str(overlay_path) if overlay_path and overlay_path.exists() else "",
                        "axis_path": str(axis_path) if axis_path and axis_path.exists() else "",
                        "pdf_path": pdf_path,
                        "manual_fill": "",
                        "manual_comment": "",
                    }
                )
    return rows


def set_link(cell, target: str, label: str) -> None:
    if not target:
        cell.value = ""
        return
    cell.value = label
    cell.hyperlink = target
    cell.style = "Hyperlink"


def save_workbook_with_locked_copy(wb: Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}__locked_copy{path.suffix}")
        wb.save(fallback)
        return fallback


def build_workbook(rows: list[dict[str, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "优先补录清单"

    headers = [
        "病例",
        "时间点",
        "页码",
        "bp_source",
        "confidence",
        "note",
        "原始PDF",
        "病例CSV",
        "监测叠图",
        "坐标对齐图",
        "人工补录值",
        "人工备注",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["case_name"])
        ws.cell(row=row_idx, column=2, value=row["time"])
        ws.cell(row=row_idx, column=3, value=row["page"])
        ws.cell(row=row_idx, column=4, value=row["bp_source"])
        ws.cell(row=row_idx, column=5, value=row["confidence"])
        ws.cell(row=row_idx, column=6, value=row["note"])
        set_link(ws.cell(row=row_idx, column=7), row["pdf_path"], "打开PDF")
        set_link(ws.cell(row=row_idx, column=8), row["csv_path"], "打开CSV")
        set_link(ws.cell(row=row_idx, column=9), row["overlay_path"], "打开叠图")
        set_link(ws.cell(row=row_idx, column=10), row["axis_path"], "打开对齐图")
        ws.cell(row=row_idx, column=11, value=row["manual_fill"])
        ws.cell(row=row_idx, column=12, value=row["manual_comment"])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {
        1: 20,
        2: 10,
        3: 8,
        4: 12,
        5: 12,
        6: 22,
        7: 12,
        8: 12,
        9: 12,
        10: 12,
        11: 24,
        12: 36,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a priority review workbook for missing trace clusters.")
    parser.add_argument(
        "--workspace",
        type=Path,
        default=ROOT,
        help="Directory containing extraction outputs and case groups. Defaults to this script directory.",
    )
    return parser.parse_args()


def configure_paths(workspace: Path) -> None:
    global ROOT, MONITOR_DIR, AXIS_DIR, OUTPUT_DIR, OUTPUT_XLSX
    ROOT = workspace
    MONITOR_DIR = ROOT / "批量监测CSV"
    AXIS_DIR = ROOT / "批量坐标对齐图"
    OUTPUT_DIR = ROOT / "人工核对包"
    OUTPUT_XLSX = OUTPUT_DIR / "优先补录清单_missing_trace_cluster.xlsx"


def main() -> None:
    args = parse_args()
    configure_paths(args.workspace.resolve())
    OUTPUT_DIR.mkdir(exist_ok=True)
    rows = collect_missing_rows()
    wb = build_workbook(rows)
    saved_path = save_workbook_with_locked_copy(wb, OUTPUT_XLSX)
    print(f"已生成: {saved_path}")
    print(f"missing trace cluster 总条目: {len(rows)}")


if __name__ == "__main__":
    main()
